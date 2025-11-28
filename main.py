import logging as nonflask_logging
from constants import *
from random import randint
from markupsafe import Markup
from flask import *
import openpyxl
import datetime
import qrcode
import bcrypt
import os

app = Flask(__name__)
app.logger.setLevel(nonflask_logging.DEBUG)

app.secret_key = os.urandom(24)

user_info_string = lambda: (f'Номер заказа: {session.get("order_number")}\nИмя: {session.get("name")}\nФамилия: {session.get("surname")}\nОтчество: {session.get("lastname")}\n'
    f'Возраст: {session.get("age")}\nАдрес электронной почты: {session.get("email")}\n'
    f'Телефонный номер: {session.get("phone")}')
purchases_converted_to_string = lambda: ''.join([f"## {v[1]} {k} ({v[0]} руб/шт) = {v[0] * v[1]} итого ## " for k, v in session.get("purchases").items()])

def update_purchases_data():
    session['total_cost'] = 0
    session['purchases_list'] = ""
    for k, v in session.get("purchases").items():
        session['purchases_list'] += f"{v[1]} {k}: {float(v[0])}\n"
        session['total_cost'] += v[0] * v[1]


@app.template_filter('nl2br')
def nl2br_filter(s):
    return Markup(s.replace('\n', '<br>\n'))


#  Рекурсия, потому что нужна обязательно по ТЗ
def test_info(name, surname, lastname, age, phone_number, email, *, entry_index=0):
    if entry_index > 4:
        if name != name.capitalize():
            session["name"] = name.capitalize()
        if surname != surname.capitalize():
            session["surname"] = surname.capitalize()
        if lastname != lastname.capitalize():
            session["lastname"] = lastname.capitalize()
        try:
            if not 0 < int(age) < 140:
                session["age"] = 1
        except:
            session["age"] = 1
        return True
    current_test = (test_name, test_surname, test_lastname, test_phone, test_email)[entry_index]
    current_arg = (name, surname, lastname, phone_number, email)[entry_index]
    if not current_test(current_arg):
        return False
    return test_info(name, surname, lastname, age, phone_number, email, entry_index=entry_index + 1)


def qr_code():
    qr_url = f"{url_for('receipt')}"
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=1,
        border=1,
    )
    qr.add_data(qr_url)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")
    qr_filename = f"qrcode.png"
    img.save(os.path.join(app.static_folder, qr_filename))
    return qr_filename


def edit_excel(*args, name=ORDERS_XLSX_PATH, method=APPEND):
    if method == APPEND:  # Добавить строку в таблицу и создать таблицу, если её не существует
        try:
            wb = openpyxl.load_workbook(name)
            ws = wb['Sheet1']
        except:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Sheet1'
        ws.append(args)
        try:
            wb.save(name)
        except Exception as e:
            app.logger.error(
                f"Таблица Excel '{name}' не была закрыта перед сохранением. Заказ не сохранён в таблицу. ({e})")
    elif method == MODIFY:  # Изменить значение в таблице по строке и столбцу на значение val (операция сложения)
        if len(args) != 3:
            app.logger.error(f"Количество аргументов для изменения значения не равно 3 (индекс строки, индекс столбца, значение)")
            return
        row, col, val = args[0], args[1], args[2]
        try:
            wb = openpyxl.load_workbook(name)
            ws = wb['Sheet1']
            if ws.cell(row, col).value is None:
                ws.cell(row, col).value = val
            else:
                ws.cell(row, col).value += val
            wb.save(name)
        except Exception as e:
            app.logger.error(f"Таблицы {name} не существует ({e})")
    elif method == CHECK:  # Проверить, что таблица существует, и добавить начальные значения, если нужно
        try:
            openpyxl.load_workbook(name)
            return True
        except:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Sheet1'
        for i in args:
            ws.append(i)
        try:
            wb.save(name)
        except Exception as e:
            app.logger.error(f"Произошла непредвиденная ошибка: {e}")
        return False
    else:
        app.logger.error(f"Несуществующий метод {method}")


def find_login_password_index(login, *, return_password=False):
    try:
        wb = openpyxl.load_workbook(f'{ACCOUNTS_XLSX_PATH}')
        ws = wb['Sheet1']
        if not return_password:
            i = 1
            for row in ws.iter_rows():
                if row[0].value == login:
                    return i
                i += 1
            return None
        else:
            for row in ws.iter_rows():
                if row[0].value == login:
                    return row[1].value
            return None
    except:
        return None


def change_ingredients(*args, allow_negative=False, enough_return=True, lack_return=False):
    if not edit_excel(name=STORAGE_XLSX_PATH, method=CHECK):
        for i in ingredient_code_transcript:
            edit_excel(ingredient_code_transcript[i], PLACEHOLDER_STORAGE_ITEM_AMOUNT, name=STORAGE_XLSX_PATH)
    wb = openpyxl.load_workbook(f'{STORAGE_XLSX_PATH}')
    ws = wb['Sheet1']
    edit_data = []  # [[row, col, val], ...]
    for arg in args:
        row_index = 1
        for row in ws.iter_rows(values_only=True):
            code = row[INGREDIENT_NAME_COLUMN_INDEX]
            for k, v in ingredient_code_transcript.items():
                if v == code:
                    code = k
                    break
            if code == arg[INGREDIENT_CODE_ING_CHAR_INDEX]:
                try:
                    if not allow_negative and row[INGREDIENT_AMOUNT_COLUMN_INDEX] - int(arg[INGREDIENT_CODE_AMT_START_INDEX:]) < 0:
                        app.logger.warning("Недостаточно ингредиентов на складе.")
                        return lack_return
                    else:
                        edit_data.append([row_index, INGREDIENT_AMOUNT_COLUMN_INDEX + 1, -int(arg[INGREDIENT_CODE_AMT_START_INDEX:])])
                except Exception as e:
                    app.logger.error(f"Не правильный синтаксис индекса ингредиента. Доп. информация: {e}")
            row_index += 1
    for data in edit_data:
        edit_excel(data[0], data[1], data[2], name=STORAGE_XLSX_PATH, method=MODIFY)
    return enough_return


@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        session['name'] = request.form.get('name')
        session['surname'] = request.form.get('surname')
        session['lastname'] = request.form.get('lastname')
        session['age'] = request.form.get('age')
        session['email'] = request.form.get('email')
        session['phone'] = request.form.get('phone')
        session['login'] = request.form.get('login')
        session['password'] = request.form.get('password')
        session['order_number'] = ['A', 'B', 'C', 'D', 'E', 'F', 'G'][randint(0, 6)] + str(randint(1, 99))
        session['custom_pizzas_amt'] = 0
        session['total_cost'] = 0
        session['purchases_list'] = ""
        session['c_pizza_cost'] = 100
        session['ingreds_in_custom_pizza'] = []
        session['purchases'] = dict()
        session['purchase_completed'] = False
        if session.get("login"):
            passwd = find_login_password_index(session.get("login"), return_password=True)
            if os.path.exists(ACCOUNTS_XLSX_PATH) and passwd:
                if not bcrypt.checkpw(session.get("password").encode("utf-8"), passwd.encode("utf-8")):
                    session["login"] = None
                    session["password"] = None
                    return f"""Введён неверный пароль<br><br><a href="{url_for('index')}">Вернуться в меню регистрации</a><br><br><a href="{url_for('main_page')}">Продолжить без аккаунта</a>"""
        if test_info(session.get("name"), session.get("surname"), session.get("lastname"), session.get("age"),
                     session.get("phone"), session.get("email")):
            return redirect(url_for('main_page'))
        else:
            return ("Введены неверные данные. Проверьте, что ФИО содержит только буквы, возраст - только цифры, "
                    "электронная почта верная, а телефонный номер соответствует формату +x-xxx-xxx-xx-xx или "
                    "x-xxx-xxx-xx-xx, где x - цифра от 0 до 9")
    return render_template("registration.html")


@app.route("/order/", methods=['GET', 'POST'])
def main_page():
    update_purchases_data()
    menu_string = f"\n{' ' * 12}МЕНЮ{' ' * 12}\n{'-' * 30}\n"
    for k, v in MENU.items():
        menu_string += f"{k}: стоимость {v}\n"
    menu_string += "-" * 30 + '\n'
    if int(session.get("age")) >= 18:
        menu_string += "МЕНЮ ДЛЯ ВЗРОСЛЫХ\n"
        for k, v in SPEC_MENU.items():
            menu_string += f"{k}: стоимость - {v} рублей\n"
        menu_string += "-" * 30
    order = f"Ваш заказ:\n{'-' * 30}\n{session.get('purchases_list')}\nИТОГО: {session.get('total_cost')} РУБЛЕЙ\n{'-' * 30}"
    return render_template("main_page.html", menu=menu_string, order_preview=order, user_info=user_info_string())


@app.route("/order/add_item", methods=["POST", "GET"])
def add_item():
    purchases = session.get('purchases')
    item = request.form.get("item_to_add", " ").strip()
    if not change_ingredients(*food_ingreds.get(item, ("-0",))):
        return f"""На складе закончились ингредиенты для этого блюда. Приносим извинения за доставленные неудобства.<br>Обратитесь за помощью к рабочему ресторана, или попробуйте заказать другое блюдо.<br><a href="{url_for('main_page')}">Вернуться в меню оформления заказа</a>"""
    try:
        purchases[item][0] += MENU[item]
        purchases[item][1] += 1
        session["purchases"] = purchases
        return redirect(url_for("main_page"))
    except:
        try:
            purchases[item][0] += SPEC_MENU[item]
            purchases[item][1] += 1
            session["purchases"] = purchases
            return redirect(url_for("main_page"))
        except:
            pass
    try:
        purchases[item] = [MENU[item], 1]
        session["purchases"] = purchases
        return redirect(url_for("main_page"))
    except:
        if int(session.get("age")) >= 18:
            try:
                purchases[item] = [SPEC_MENU[item], 1]
                session["purchases"] = purchases
                return redirect(url_for("main_page"))
            except:
                return f"""В меню нет предмета {item}<br><a href="{url_for('main_page')}">Вернуться в меню оформления заказа</a>"""
        else:
            return f"""В детском меню нет предмета {item}<br><a href="{url_for('main_page')}">Вернуться в меню оформления заказа</a>"""


@app.route("/order/custom/modify_custom_pizzas_amt", methods=["POST", "GET"])
def increase_custom_pizza_counter():
    session['custom_pizzas_amt'] += 1
    return redirect(url_for('custom_pizza'))


@app.route("/order/custom/", methods=["POST", "GET"])
def custom_pizza():
    ingreds = "-" * 30 + '\n'
    for k, v in INGREDIENTS.items():
        ingreds += f"{k}: стоимость {v}\n"
    ingreds_in_pizza = " - Тесто (100 рублей)\n"
    for i in session.get("ingreds_in_custom_pizza"):
        ingreds_in_pizza += f" - {i} ({INGREDIENTS[i]} рублей)\n"
        session['c_pizza_cost'] += INGREDIENTS[i]
    return render_template("custom_pizza.html", ingreds_in_pizza=ingreds_in_pizza, ingredients=ingreds, user_info=user_info_string())


@app.route("/order/custom/add_ingredient", methods=["POST", "GET"])
def custom_pizza_add_ingredient():
    ingred = request.form.get("ingredient", " ").strip()
    if not change_ingredients(*food_ingreds.get(ingred, ("-0",))):
        return f"""На складе закончился этот ингредиент. Приносим извинения за доставленные неудобства.<br>Обратитесь за помощью к рабочему ресторана, или попробуйте добавить другой ингредиент.<br><a href="{url_for('custom_pizza')}">Вернуться в меню оформления заказа</a>"""
    try:
        session['c_pizza_cost'] += INGREDIENTS[ingred]
        session['ingreds_in_custom_pizza'].append(ingred)
        return redirect(url_for("custom_pizza"))
    except:
        return f"""Ингредиента {ingred} не существует<br><a href="{url_for('custom_pizza')}">Вернуться в меню создания кастомной пиццы</a>"""


@app.route("/order/custom/finish_custom_pizza", methods=["POST", "GET"])
def finish_custom_pizza():
    session["purchases"][f"Кастомная пицца {session.get('custom_pizzas_amt')}"] = [session.get('c_pizza_cost'), 1]
    session['c_pizza_cost'] = 100
    session['ingreds_in_custom_pizza'] = []
    return redirect(url_for("main_page"))


@app.route("/order/custom/decline", methods=["POST", "GET"])
def decline_custom_pizza():
    session['c_pizza_cost'] = 100
    session['ingreds_in_custom_pizza'] = []
    return redirect(url_for("main_page"))


@app.route("/order/pay/", methods=["POST", "GET"])
def pay():
    update_purchases_data()
    order = f"Ваш заказ:\n{'-' * 30}\n{session.get('purchases_list')}\nИТОГО: {session.get('total_cost')} РУБЛЕЙ\n{'-' * 30}"
    return render_template("pay.html", order_number=session.get("order_number"), purchases=order, total_cost=session.get('total_cost'), user_info=user_info_string())


@app.route("/order/pay/card", methods=["POST", "GET"])
def pay_card():
    bonus_info = "Вам не были начислены бонусы, так как вы не зарегистрированы"
    if session.get("login") and session.get("password"):
        if not session.get('purchase_completed'):
            edit_excel(session.get("order_number"), user_info_string(), session.get("total_cost"),
                       purchases_converted_to_string(),
                       session.get("login"), "Оплачено картой")
            passwd_i = find_login_password_index(session.get("login"))
            if passwd_i:
                edit_excel(passwd_i, 3, session.get('total_cost') // 10, name=ACCOUNTS_XLSX_PATH, method=MODIFY)
            else:
                edit_excel(session.get("login"),
                           bcrypt.hashpw(session.get("password").encode("utf-8"), bcrypt.gensalt()).decode(),
                           session.get('total_cost') // 10, name=ACCOUNTS_XLSX_PATH)
        bonus_info = f"На ваш аккаунт начислено {session.get('total_cost') // 10} бонусов"
    else:
        if not session.get('purchase_completed'):
            edit_excel(session.get("order_number"), user_info_string(), session.get("total_cost"), purchases_converted_to_string(),
                       "Оплачено картой")
    pay_info = "Оплата: картой"
    session['purchase_completed'] = True
    return render_template("thanks_for_the_purchase.html", order_number=session.get('order_number'), bonus_info=bonus_info, user_info=user_info_string(), pay_info=pay_info)


@app.route("/order/pay/cash", methods=["POST", "GET"])
def pay_cash():
    session["cash"] = request.form.get("cash")
    app.logger.debug(session.get("cash"))
    cash = 0
    try:
        cash = int(session.get("cash"))
    except:
        pass
    if cash and cash - int(session.get("total_cost")) >= 0:
        bonus_info = "Вам не были начислены бонусы, так как вы не зарегистрированы"
        if session.get("login") and session.get("password"):
            if not session.get('purchase_completed'):
                edit_excel(session.get("order_number"), user_info_string(), session.get("total_cost"),
                           purchases_converted_to_string(),
                           session.get("login"), "Оплачено наличными")
                passwd_i = find_login_password_index(session.get("login"))
                if passwd_i:
                    edit_excel(passwd_i, 3, session.get('total_cost') // 10, name=ACCOUNTS_XLSX_PATH, method=MODIFY)
                else:
                    edit_excel(session.get("login"),
                               bcrypt.hashpw(session.get("password").encode("utf-8"), bcrypt.gensalt()).decode(),
                               session.get('total_cost') // 10, name=ACCOUNTS_XLSX_PATH)
            bonus_info = f"На ваш аккаунт начислено {session.get('total_cost') // 10} бонусов"
        else:
            if not session.get('purchase_completed'):
                edit_excel(session.get("order_number"), user_info_string(), session.get("total_cost"),
                           purchases_converted_to_string(), "Оплачено наличными")
        pay_info = f"Оплата: наличными\nСдача: {cash - session.get('total_cost')}"
        session['purchase_completed'] = True
        return render_template("thanks_for_the_purchase.html", order_number=session.get('order_number'),
                               bonus_info=bonus_info, user_info=user_info_string(), pay_info=pay_info)
    else:
        return redirect(url_for("pay"))


@app.route("/order/pay/receipt", methods=["POST", "GET"])
def receipt():
    update_purchases_data()
    return render_template("receipt.html", date=datetime.date.today(), order_number=session.get("order_number"), client=user_info_string(), total_cost=session.get("total_cost"), purchases_list=session.get("purchases_list"), qr_url=url_for('static', filename=qr_code()))


@app.route("/order/reset", methods=["POST", "GET"])
def reset_data():
    session['order_number'] = ['A', 'B', 'C', 'D', 'E', 'F', 'G'][randint(0, 6)] + str(randint(1, 99))
    session['custom_pizzas_amt'] = 0
    session['total_cost'] = 0
    session['purchases_list'] = ""
    session['c_pizza_cost'] = 100
    session['ingreds_in_custom_pizza'] = []
    session['purchases'] = dict()
    session['purchase_completed'] = False
    return redirect(url_for("main_page"))

